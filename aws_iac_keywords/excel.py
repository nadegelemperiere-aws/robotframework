""" -----------------------------------------------------
# Copyright (c) [2022] Nadege Lemperiere
# All rights reserved
# -------------------------------------------------------
# Keywords to manage excel file content testing
# -------------------------------------------------------
# Nadège LEMPERIERE, @05 october 2021
# Latest revision: 05 october 2021
# --------------------------------------------------- """

# Openpyxl includes
from openpyxl import load_workbook

# Robotframework includes
from robot.api import logger
from robot.api.deco import keyword
ROBOT = False

@keyword("Cell Content Shall Match")
def check_cell_content(filename, sheet, identifier, column, value) :
    """ Test if an excel file test content match the hypothesis
        ---
        filename  (str) : Xlsx filename to analyze
        sheet     (str) : Excel sheet in which the cell is located
        identifier(str) : identifier identifying the line to analyse (in the Identifier column)
        column    (str) : Header of the column to analyze
        value     (str) : Value to look for in the selected cell
    """

    full_filename = filename

    # Load workbook
    wbook = load_workbook(full_filename)

    # Select sheet
    content_sheet = wbook[sheet]

    # Associate header to column
    i_column = 1
    column_to_header = {}
    header_to_column = {}
    content = content_sheet.cell(1,i_column).value
    while content is not None :
        column_to_header[i_column]  = content
        header_to_column[content]   = i_column
        i_column = i_column + 1
        content = content_sheet.cell(1,i_column).value
    if not column in header_to_column : raise Exception('Column not found')

	# Look for the line in which identifier is located
    selected_row = -1
    for i_row in range(1,content_sheet.max_row + 1) :
        logger.info(content_sheet.cell(i_row,header_to_column['Identifier']).value)
        logger.info(identifier)
        if str(content_sheet.cell(i_row,header_to_column['Identifier']).value) == identifier :
            logger.info("found")
            selected_row = i_row

    if selected_row == -1 : raise Exception('Identifier not found')

    if not content_sheet.cell(selected_row,header_to_column[column]).value == value :
        raise Exception('Value does not match')
